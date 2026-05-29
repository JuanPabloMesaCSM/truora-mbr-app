# -*- coding: utf-8 -*-
"""
Auditoría: compara la fuente de verdad (Excel "Oppy NRR 2026" hoja
"Clientes Oppy 2026") vs la tabla `clientes` en Supabase.

Reporta:
  1. Clientes en Excel que NO existen en Supabase (faltantes).
  2. Clientes en Supabase con TCI distinto al del Excel (mismatch TCI).
  3. Clientes en Supabase con CSM distinto al del Excel (mismatch CSM).
  4. Clientes en Supabase con TCI que NO aparece en Excel (huérfanos).
  5. Estado Activo/Inactivo desalineado.

REQUIERE env var SUPABASE_SERVICE_ROLE_KEY.

Uso (Git Bash):
  export SUPABASE_SERVICE_ROLE_KEY=$(grep SUPABASE_SERVICE_ROLE_KEY ../../truora-mbr-app/.env.local | cut -d= -f2)
  python audit.py
"""

import json
import os
import re
import sys
import urllib.request
import warnings
from collections import defaultdict
from pathlib import Path

warnings.filterwarnings("ignore")

import openpyxl

ROOT = Path(__file__).resolve().parent
EXCEL_PATH = Path("C:/Users/Administrador/Downloads/Oppy NRR 2026 (1).xlsx")
SHEET_NAME = "Clientes Oppy 2026"

SUPABASE_URL = "https://cjrhxmfnmajxiwiiuwym.supabase.co"
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
if not SUPABASE_KEY:
    sys.exit("ERROR: falta env var SUPABASE_SERVICE_ROLE_KEY")

ADMIN_EMAILS = {"amarquez@truora.com", "jdiaz@truora.com"}

# Mapeo nombre CSM (Excel) → email (Supabase). Construido del SELECT de csm.
CSM_NAME_TO_EMAIL = {
    "ana milena marquez": "amarquez@truora.com",
    "ana milena márquez": "amarquez@truora.com",
    "ana milena": "amarquez@truora.com",
    "anami": "amarquez@truora.com",
    "ana": "amarquez@truora.com",
    "daniela tibaquira": "dtibaquira@truora.com",
    "daniela tibaquirá": "dtibaquira@truora.com",
    "daniela": "dtibaquira@truora.com",
    "elisa varela": "evarela@truora.com",
    "elisa": "evarela@truora.com",
    "juan pablo diaz": "jdiaz@truora.com",
    "juan pablo díaz": "jdiaz@truora.com",
    "jdiaz": "jdiaz@truora.com",
    "jp diaz": "jdiaz@truora.com",
    "juan pablo mesa": "jpmesa@truora.com",
    "jp mesa": "jpmesa@truora.com",
    "jpmesa": "jpmesa@truora.com",
    "juan pablo otoya": "jpotoya@truora.com",
    "jp otoya": "jpotoya@truora.com",
    "japo": "jpotoya@truora.com",
    "jpotoya": "jpotoya@truora.com",
    "natalia gutierrez": "nagutierrez@truora.com",
    "natalia gutiérrez": "nagutierrez@truora.com",
    "natalia": "nagutierrez@truora.com",
    "sebastian duran": "sduran@truora.com",
    "sebastián durán": "sduran@truora.com",
    "sebas": "sduran@truora.com",
    "soporte": "soporte@truora.com",
    "valentina arango": "varango@truora.com",
    "valentina": "varango@truora.com",
    "valeria lopez": "vlopez@truora.com",
    "valeria lópez": "vlopez@truora.com",
    "valeria": "vlopez@truora.com",
}


def supabase_get(path: str):
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/{path}",
        headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"},
    )
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())


def normalize_name(s: str | None) -> str:
    if not s:
        return ""
    return re.sub(r"\s+", " ", s).strip().lower()


TCI_RE = re.compile(r"\bTCI[a-f0-9]{32}\b", re.IGNORECASE)


def extract_tci(s: str | None) -> str | None:
    """Extrae el TCI de un string que puede tener descripción extra (ej:
    'TCIxxx - Dev test')."""
    if not s:
        return None
    s = str(s).strip()
    m = TCI_RE.search(s)
    if m:
        return m.group(0)
    # Fallback: algunos client_id no son TCI (ej: '5aurr2fgj3...').
    if len(s) > 8 and " " not in s and "-" not in s.split()[0]:
        return s.split()[0]
    return None


def csm_to_email(csm_name: str | None) -> str | None:
    n = normalize_name(csm_name)
    if not n:
        return None
    return CSM_NAME_TO_EMAIL.get(n)


def main():
    # ---- Excel ----
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    excel_rows = []  # list of dicts
    for row in range(2, ws.max_row + 1):
        nombre = ws.cell(row, 1).value
        client_id_raw = ws.cell(row, 2).value
        csm_name = ws.cell(row, 3).value
        estado = ws.cell(row, 4).value

        if not nombre:
            continue

        tci = extract_tci(client_id_raw)
        csm_email = csm_to_email(csm_name)

        excel_rows.append({
            "row": row,
            "nombre": str(nombre).strip(),
            "tci": tci,
            "tci_raw": str(client_id_raw or "").strip(),
            "csm_name": str(csm_name or "").strip(),
            "csm_email": csm_email,
            "estado": str(estado or "").strip(),
        })

    print(f"=== Excel: {len(excel_rows)} filas leídas ===")

    # Filtrar solo Activos para el análisis principal
    excel_active = [r for r in excel_rows if r["estado"].lower() == "activo"]
    excel_inactive = [r for r in excel_rows if r["estado"].lower() != "activo"]
    print(f"  Activos: {len(excel_active)}")
    print(f"  Otros estados: {len(excel_inactive)}")
    print()

    # ---- Supabase ----
    sb_rows = supabase_get("clientes?select=id,nombre,csm_email,activo,client_id_di,client_id_bgc,client_id_ce&limit=2000")
    print(f"=== Supabase: {len(sb_rows)} filas en `clientes` ===")
    sb_real = [r for r in sb_rows if r["csm_email"] not in ADMIN_EMAILS]
    print(f"  Sin admin-duplicates (CSM real): {len(sb_real)}")
    print()

    # Index Supabase (CSM real) por TCI: TCI -> fila
    by_tci = {}
    for r in sb_real:
        for col in ("client_id_di", "client_id_bgc", "client_id_ce"):
            tci = r[col]
            if tci:
                by_tci.setdefault(tci, []).append({"col": col, "row": r})

    # Index Supabase por nombre normalizado
    by_name = defaultdict(list)
    for r in sb_real:
        by_name[normalize_name(r["nombre"])].append(r)

    # ============================================================================
    # ANÁLISIS
    # ============================================================================

    print("\n" + "=" * 80)
    print("DIAGNÓSTICO: Excel (Activos) vs Supabase (CSM real)")
    print("=" * 80)

    not_in_supabase = []     # Excel tiene pero Supabase no
    tci_mismatch = []        # nombre matchea pero TCI distinto
    csm_mismatch = []        # TCI matchea pero CSM distinto
    csm_no_resuelto = []     # csm en Excel no se pudo mapear a email
    no_tci_excel = []        # Excel sin TCI parseable
    inactivo_supabase = []   # Activo en Excel pero inactivo en Supabase

    matched_tcis = set()     # TCIs del Excel que sí matchearon en Supabase

    for r in excel_active:
        if not r["tci"]:
            no_tci_excel.append(r)
            continue
        if not r["csm_email"]:
            csm_no_resuelto.append(r)

        sb_match_by_tci = by_tci.get(r["tci"])
        if sb_match_by_tci:
            matched_tcis.add(r["tci"])
            sb_row = sb_match_by_tci[0]["row"]
            # Verificar CSM
            if r["csm_email"] and sb_row["csm_email"] != r["csm_email"]:
                csm_mismatch.append({
                    "excel": r,
                    "sb": sb_row,
                })
            # Verificar activo
            if not sb_row["activo"]:
                inactivo_supabase.append({"excel": r, "sb": sb_row})
        else:
            # No match por TCI. Buscar por nombre.
            sb_match_by_name = by_name.get(normalize_name(r["nombre"]))
            if sb_match_by_name:
                tci_mismatch.append({
                    "excel": r,
                    "sb": sb_match_by_name[0],
                })
            else:
                not_in_supabase.append(r)

    # Supabase TCIs que NO están en Excel
    huerfanos = []
    for r in sb_real:
        for col in ("client_id_di", "client_id_bgc", "client_id_ce"):
            tci = r[col]
            if tci and tci not in matched_tcis:
                huerfanos.append({"row": r, "tci": tci, "col": col})

    # ============================================================================
    # REPORTE
    # ============================================================================

    def section(title, rows, fmt):
        print(f"\n--- {title} ({len(rows)}) ---")
        for r in rows:
            print(f"  {fmt(r)}")

    section(
        "[!] EN EXCEL ACTIVO PERO NO EN SUPABASE",
        not_in_supabase,
        lambda r: f"{r['nombre']:30} | TCI={r['tci']} | CSM={r['csm_name']}",
    )

    section(
        "[*] TCI MISMATCH (mismo nombre, TCI distinto)",
        tci_mismatch,
        lambda r: f"{r['excel']['nombre']:30} | Excel={r['excel']['tci']} | Supabase=DI:{r['sb']['client_id_di']} BGC:{r['sb']['client_id_bgc']} CE:{r['sb']['client_id_ce']}",
    )

    section(
        "[*] CSM MISMATCH (TCI matchea pero CSM distinto)",
        csm_mismatch,
        lambda r: f"{r['excel']['nombre']:30} | TCI={r['excel']['tci']} | Excel={r['excel']['csm_email']} | Supabase={r['sb']['csm_email']}",
    )

    section(
        "[?] CSM NO RESUELTO (nombre Excel no mapea a email)",
        csm_no_resuelto,
        lambda r: f"{r['nombre']:30} | CSM Excel='{r['csm_name']}'",
    )

    section(
        "[?] EXCEL SIN TCI PARSEABLE",
        no_tci_excel,
        lambda r: f"{r['nombre']:30} | Client ID raw='{r['tci_raw']}' | CSM={r['csm_name']}",
    )

    section(
        "[~] ACTIVO EN EXCEL PERO INACTIVO EN SUPABASE",
        inactivo_supabase,
        lambda r: f"{r['excel']['nombre']:30} | TCI={r['excel']['tci']}",
    )

    print(f"\n--- [-] TCIs EN SUPABASE NO PRESENTES EN EXCEL ACTIVO ({len(huerfanos)}) ---")
    seen = set()
    for h in huerfanos:
        k = (h["row"]["nombre"], h["tci"])
        if k in seen:
            continue
        seen.add(k)
        print(f"  {h['row']['nombre']:30} | TCI={h['tci']:40} | col={h['col']:13} | CSM={h['row']['csm_email']}")

    print("\n" + "=" * 80)
    print("RESUMEN")
    print("=" * 80)
    print(f"  [!] Faltantes en Supabase:        {len(not_in_supabase)}")
    print(f"  [*] TCI mismatch:                 {len(tci_mismatch)}")
    print(f"  [*] CSM mismatch:                 {len(csm_mismatch)}")
    print(f"  [?] CSM no resuelto en Excel:     {len(csm_no_resuelto)}")
    print(f"  [?] Excel sin TCI parseable:      {len(no_tci_excel)}")
    print(f"  [~] Activos vs inactivos:         {len(inactivo_supabase)}")
    seen_h = set()
    for h in huerfanos:
        seen_h.add((h["row"]["nombre"], h["tci"]))
    print(f"  [-] Huerfanos en Supabase:        {len(seen_h)} TCIs unicos")


if __name__ == "__main__":
    main()
