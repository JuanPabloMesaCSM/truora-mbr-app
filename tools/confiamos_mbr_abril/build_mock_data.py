# -*- coding: utf-8 -*-
"""
Genera el JSON con shape DI esperado por SlideCanvas para Confiamos abril 2026.

Output: confiamos_abril_mock.json — pegar en MockConfiamosAbril.tsx.

El JSON imita el payload del workflow Report Builder DI pero construido a partir
del CSV de validations standalone (sin identity_process_id). Algunas columnas
no aplicables se ponen en "0" para que los slides rendericen sin warnings.
"""

import csv
import json
import warnings
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

warnings.filterwarnings("ignore")

import openpyxl

ROOT = Path(__file__).resolve().parent
ABRIL_CSV = Path("C:/Users/Administrador/Downloads/HRP09b8bc329cd86360575934b327e1ea1f.csv")
MARZO_XLSX = Path("C:/Users/Administrador/Downloads/Consumos por validador Marzo 2026.xlsx")
OUTPUT = ROOT / "confiamos_abril_mock.json"


def parse_iso(s):
    if not s:
        return None
    s = str(s).strip()
    try:
        return datetime.strptime(s, "%Y-%m-%dT%H:%M:%SZ")
    except ValueError:
        try:
            return datetime.fromisoformat(s.replace("Z", "+00:00"))
        except Exception:
            return None


def normalize_status(s):
    s = (s or "").strip().lower()
    if s in ("exitoso", "success"):
        return "exitoso"
    if s in ("fallido", "failed", "rejected"):
        return "fallido"
    return s or "sin_estado"


def normalize_motivo(motivo, estado):
    m = (motivo or "").strip()
    e = (estado or "").strip()
    return m or e or "sin_motivo"


def to_snake(s):
    """'similarity is below threshold' -> 'similarity_is_below_threshold'."""
    s = (s or "").strip().lower()
    s = "".join(c if (c.isalnum() or c == "_") else "_" for c in s)
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")[:60] or "sin_motivo"


def read_csv_rows(path):
    rows = []
    with open(path, encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f, delimiter=";")
        for r in reader:
            rows.append(r)
    return rows


def read_xlsx_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        rd = {h: ws.cell(r, c).value for c, h in enumerate(headers, start=1)}
        rows.append(rd)
    return rows


def analyze(rows):
    total = len(rows)
    by_status = Counter(normalize_status(r.get("validation_status")) for r in rows)
    exitosos = by_status.get("exitoso", 0)
    fallidos = by_status.get("fallido", 0)

    # Por tipo
    by_type_status = defaultdict(lambda: Counter())
    by_type_motivos = defaultdict(Counter)
    for r in rows:
        t = (r.get("type") or "").strip()
        st = normalize_status(r.get("validation_status"))
        by_type_status[t][st] += 1
        if st == "fallido":
            by_type_motivos[t][normalize_motivo(r.get("Motivo de rechazo"), r.get("Estado de falla"))] += 1

    # Top motivos global
    all_motivos = Counter()
    for r in rows:
        if normalize_status(r.get("validation_status")) == "fallido":
            all_motivos[normalize_motivo(r.get("Motivo de rechazo"), r.get("Estado de falla"))] += 1

    # Usuarios únicos por IP (proxy)
    by_ip = Counter()
    by_ip_status = defaultdict(Counter)
    for r in rows:
        ip = r.get("ip_address") or "(sin ip)"
        by_ip[ip] += 1
        by_ip_status[ip][normalize_status(r.get("validation_status"))] += 1

    usuarios_unicos = len(by_ip)
    usuarios_exitosos = sum(1 for ip, c in by_ip_status.items() if c.get("exitoso", 0) > 0)

    # Distribución de intentos por usuario
    intentos_dist = Counter()
    for ip, c in by_ip.items():
        if c == 1:
            intentos_dist[1] += 1
        elif c == 2:
            intentos_dist[2] += 1
        elif c == 3:
            intentos_dist[3] += 1
        else:
            intentos_dist["4+"] += 1

    return {
        "total": total,
        "exitosos": exitosos,
        "fallidos": fallidos,
        "tasa_exito": round(100.0 * exitosos / total, 1) if total else 0,
        "by_type_status": dict(by_type_status),
        "by_type_motivos": dict(by_type_motivos),
        "all_motivos": all_motivos,
        "usuarios_unicos": usuarios_unicos,
        "usuarios_exitosos": usuarios_exitosos,
        "intentos_dist": intentos_dist,
        "ratio_intentos": round(total / usuarios_unicos, 2) if usuarios_unicos else 0,
    }


def s(n):
    """JSON string de un entero/float (Report Builder espera strings)."""
    if isinstance(n, float):
        return str(round(n, 1))
    return str(int(n))


def build_payload(abr, mar):
    """Construye el JSON con shape Report Builder DI."""
    doc_st = abr["by_type_status"].get("document-validation", Counter())
    doc_total = sum(doc_st.values())
    doc_exito = doc_st.get("exitoso", 0)
    doc_falla = doc_st.get("fallido", 0)
    doc_pct = round(100.0 * doc_exito / doc_total, 1) if doc_total else 0

    mar_doc_st = mar["by_type_status"].get("document-validation", Counter())
    mar_doc_total = sum(mar_doc_st.values())
    mar_doc_exito = mar_doc_st.get("exitoso", 0)
    mar_doc_pct = round(100.0 * mar_doc_exito / mar_doc_total, 1) if mar_doc_total else 0

    face_st = abr["by_type_status"].get("face-recognition", Counter())
    face_total = sum(face_st.values())
    face_exito = face_st.get("exitoso", 0)
    face_pct = round(100.0 * face_exito / face_total, 1) if face_total else 0

    mar_face_st = mar["by_type_status"].get("face-recognition", Counter())
    mar_face_total = sum(mar_face_st.values())
    mar_face_exito = mar_face_st.get("exitoso", 0)
    mar_face_pct = round(100.0 * mar_face_exito / mar_face_total, 1) if mar_face_total else 0

    # Variación porcentual abril vs marzo (volumen)
    var_pct = round(100.0 * (abr["total"] - mar["total"]) / mar["total"], 1) if mar["total"] else 0

    # ---- 1_metricas_generales ----
    # col1=total, col2=exitosos, col3=fallidos, col4=declinados, col5=expirados,
    # col6=errores_tec, col7=cancelados, col8=conv_pct,
    # col9=total_prev, col10=exitosos_prev, col11=conv_prev, col_extra1=variacion_volumen
    metricas_gen = {
        "bloque": "1_metricas_generales",
        "col1": s(abr["total"]),
        "col2": s(abr["exitosos"]),
        "col3": s(abr["fallidos"]),  # fallidos totales (sin diferenciar declinados)
        "col4": s(abr["fallidos"]),  # declinados = fallidos en este modelo
        "col5": "0",                  # expirados (no aplica en standalone)
        "col6": "0",                  # errores tec
        "col7": "0",                  # cancelados
        "col8": s(abr["tasa_exito"]),
        "col9": s(mar["total"]),
        "col10": s(mar["exitosos"]),
        "col11": s(mar["tasa_exito"]),
        "col_extra1": str(round(var_pct, 1)),
    }

    # ---- 2_usuarios_reintentos ----
    # OMITIDO para Confiamos: el cliente llama desde backend con IP única,
    # por eso "usuarios únicos" y "reintentos" no son métricas significativas
    # bajo el modelo de validations standalone.

    # ---- 3_validaciones_doc_rostro ----
    # col1=doc_total, col2=doc_exitosos, col3=doc_pct
    # col4=doc_total_prev, col5=doc_pct_prev
    # col6=rostro_total, col7=rostro_exitosos, col8=rostro_pct
    # col9=rostro_total_prev, col10=rostro_pct_prev
    # col11=doc_rechazados, col_extra1=doc_rechazados_prev
    # col_extra2=rostro_rechazados, col_extra3=rostro_rechazados_prev
    val_doc_rostro = {
        "bloque": "3_validaciones_doc_rostro",
        "col1": s(doc_total), "col2": s(doc_exito), "col3": s(doc_pct),
        "col4": s(mar_doc_total), "col5": s(mar_doc_pct),
        "col6": s(face_total), "col7": s(face_exito), "col8": s(face_pct),
        "col9": s(mar_face_total), "col10": s(mar_face_pct),
        "col11": s(doc_falla),
        "col_extra1": s(mar_doc_st.get("fallido", 0)),
        "col_extra2": s(face_st.get("fallido", 0)),
        "col_extra3": s(mar_face_st.get("fallido", 0)),
    }

    # ---- 4_historico_3meses ----
    # 1 fila por mes con periodo, col1=total, col2=exitosos, col3=conv_pct,
    # col4=usuarios_unicos, col5=conv_por_usuario_pct
    # Solo tenemos marzo + abril (no enero/febrero).
    # col4/col5 (usuarios) no aplica para Confiamos — IP única backend, ponemos 0.
    historico = [
        {
            "bloque": "4_historico_3meses",
            "periodo": "2026-03-01",
            "col1": s(mar["total"]), "col2": s(mar["exitosos"]), "col3": s(mar["tasa_exito"]),
            "col4": "0", "col5": "0",
        },
        {
            "bloque": "4_historico_3meses",
            "periodo": "2026-04-01",
            "col1": s(abr["total"]), "col2": s(abr["exitosos"]), "col3": s(abr["tasa_exito"]),
            "col4": "0", "col5": "0",
        },
    ]

    # ---- 7_razones_doc ----
    razones_doc = []
    for motivo, count in abr["by_type_motivos"].get("document-validation", Counter()).most_common(8):
        razones_doc.append({"bloque": "7_razones_doc", "col1": to_snake(motivo), "col2": s(count)})

    # ---- 8_razones_rostro ----
    razones_rostro = []
    for motivo, count in abr["by_type_motivos"].get("face-recognition", Counter()).most_common(8):
        razones_rostro.append({"bloque": "8_razones_rostro", "col1": to_snake(motivo), "col2": s(count)})

    # ---- 10_declinados (top global de fallidos) ----
    declinados = []
    for motivo, count in abr["all_motivos"].most_common(10):
        declinados.append({"bloque": "10_declinados", "col1": to_snake(motivo), "col2": s(count)})

    payload = {
        "1_metricas_generales": [metricas_gen],
        "3_validaciones_doc_rostro": [val_doc_rostro],
        "4_historico_3meses": historico,
        "7_razones_doc": razones_doc,
        "8_razones_rostro": razones_rostro,
        "10_declinados": declinados,
    }
    return payload


def main():
    print("Leyendo CSVs...")
    abr_rows = read_csv_rows(ABRIL_CSV)
    mar_rows = read_xlsx_rows(MARZO_XLSX)

    abr = analyze(abr_rows)
    mar = analyze(mar_rows)

    print(f"Marzo: {mar['total']} validaciones, {mar['usuarios_unicos']} IPs únicas")
    print(f"Abril: {abr['total']} validaciones, {abr['usuarios_unicos']} IPs únicas")

    payload = build_payload(abr, mar)

    OUTPUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\nOK: {OUTPUT}")
    print(f"Tamano: {OUTPUT.stat().st_size} bytes")


if __name__ == "__main__":
    main()
